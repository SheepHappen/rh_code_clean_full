import csv
import io
import requests
import os
from os import walk
import zipfile

from openpyxl import load_workbook

from core.settings import SASB_API_KEY
from django.core.management.base import BaseCommand, CommandError
from django.db.models import Q

from company.models import Industry, SasbCompanyLookup

from pprint import pprint

def download_sasb_data():
    url = "https://licensing.sasb.org/api/v1.0/getProduct/PL-SIC"
    headers = {
        'Authorization': SASB_API_KEY
    }
    req = requests.post(url, headers=headers)
    data_file = None
    if req.status_code == 200:
        try:
            content = zipfile.ZipFile(io.BytesIO(req.content))
            print('has content', content)
            for name in content.namelist():
                if 'final_new_format.xlsx' or '_final.xlsx' in name:
                    data_file = content.read(name)
        except:
            pass

    return data_file, req.status_code

def open_file(data_file):
    return load_workbook(filename=io.BytesIO(data_file))


def parse_file(data_file):
    sheet = data_file.active
    for row in sheet.iter_rows(min_row=2):
        industry = Industry.objects.filter(
            Q(name=row[5].value) | Q(name=row[7].value)
        )
        if industry:
            SasbCompanyLookup.objects.get_or_create(
                name=row[1].value,
                industry=industry[0]
            )


def parse_file_for_specific_industry(data_file, industry_name):
    sheet = data_file.active
    for row in sheet.iter_rows(min_row=2):
        if industry_name in [row[5].value, row[7].value]:
            industry = Industry.objects.filter(
                Q(name=row[5].value) | Q(name=row[7].value)
            )
            if industry:
                SasbCompanyLookup.objects.get_or_create(
                    name=row[1].value,
                    industry=industry[0]
                )


class Command(BaseCommand):
    def add_arguments(self, parser):
        # Named (optional) arguments
        parser.add_argument('--industry_name', type=str)
        parser.add_argument('--update_all', type=str)

    def handle(self, *args, **options):
        industry_name = options.get('industry_name', None)
        update_all = options.get('update_all', False)

        if not industry_name and not update_all:
            SasbCompanyLookup.objects.all().delete()

        data_file, status = download_sasb_data()

        if data_file is not None and status == 200:
            data_file = open_file(data_file)

            if not industry_name:
                parse_file(data_file)
            else:
                parse_file_for_specific_industry(data_file, industry_name)

        return "{}".format(status)