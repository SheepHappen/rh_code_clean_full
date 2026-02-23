import company
import io
from datetime import datetime
import tablib
import json
import uuid
from copy import copy
from collections import ChainMap

from openpyxl import Workbook, styles
from openpyxl.writer.excel import save_virtual_workbook

from requests.auth import HTTPBasicAuth
import requests

from django.views.generic import View
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.template.defaultfilters import slugify

from country.models import Country

from core.settings import ASSESSMENT_SESSION_TIME
from core.models import DocumentQuestion, ManagementQuestion, Threshold
from core.models import Permission
from core.utils import get_category_color, get_score_colour, is_ajax
from .models import (
    AssessmentIssueRecommendations,
    CompanyAssessment,
    Industry,
    DocumentQuestionAnswerSet,
    ManagementQuestionAnswerSet,
    Company,
    IndustryType,
    SasbCompanyLookup
)
from .calc import calculate_rev_counter_score, calculate_top_5_at_risk_countrys, calculate_top_5_impact_score, calculate_document_question_complete
from .forms import KeyDetailsForm, CompanyFootprintForm, RecommendationForm, CompanyIndustryForm
from riskfactor.models import (
    IndustryRiskDataSet,
    MaterialityRisk,
    RiskDataSet,
    CountryRisk
)


def get_assessment(slug, company):
    return CompanyAssessment.objects.select_related(
        'created_by'
    ).select_related(
        'company_fund'
    ).select_related('custom_risk_score').prefetch_related('industries').prefetch_related('countries').get(
        slug=slug,
        created_by__userprofile__company=company,
    )

# def create_new_version(assessment):
#     now = datetime.now()
#     time_diff = now - assessment.date_created.replace(tzinfo=None)
#     time_in_mins = time_diff.seconds // 60 % 60
#     if time_in_mins > int(ASSESSMENT_SESSION_TIME):
#         assessment.is_latest = False
#         assessment.save()

#         new_assessment = copy(assessment)
#         new_assessment.pk = None
#         new_assessment.id = None
#         new_assessment.parent = None
#         new_assessment.is_latest = True
#         new_assessment.save()

#         assessment.parent = new_assessment
#         assessment.save()

#         children = CompanyAssessment.objects.filter(parent=assessment)

#         for child in children:
#             child.parent = new_assessment
#             child.is_latest = False
#             child.save()

#         return new_assessment
#     else:
#         return assessment


def leading_zeros(value, desired_digits):
    """
    Given an integer, returns a string representation, padded with [desired_digits] zeros.
    """
    num_zeros = int(desired_digits) - len(str(value))
    padded_value = []
    while num_zeros >= 1:
        padded_value.append("0")
        num_zeros = num_zeros - 1
    padded_value.append(str(value))
    return "".join(padded_value)


def get_industry_types(assessment):
    industry_types = []
    assessment_industries = assessment.industries.all()
    for industry_type in IndustryType.objects.all().prefetch_related('industry_set'):
        options = []
        count = 0
        for option in industry_type.industry_set.all():
            if option in assessment_industries:
                count += 1
                option.checked = True
            options.append(option)

        industry_types.append({
            'industry_type': industry_type,
            'options': options,
            'count': count
        })

    return assessment_industries.count(), industry_types


def sasb_get_industry(assessment):
    if 'SASB' in assessment.company_name:
        sasb_company = SasbCompanyLookup.objects.filter(
            name__contains=assessment.company_name.replace('SASB:', '').strip()
        ).select_related('industry')
        if sasb_company:
            return sasb_company.first().industry
    return None


def get_document_question_answer(assessment, question):
    answer = DocumentQuestionAnswerSet.objects.filter(
        assessment=assessment,
        question=question
    ).select_related('assessment').select_related('question')
    if answer:
        return answer.first()


def get_key_document_answer_count(assessment):
    key_policy_questions = DocumentQuestion.objects.filter(
        Q(is_key=True) | Q(is_best_practice=True) | Q(is_market_leading=True)
    )
    key_policy_answers = DocumentQuestionAnswerSet.objects.filter(
        assessment=assessment, question__in=key_policy_questions
    ).filter(
        Q(answer='Y') & Q(question__yes_score__gt=0)
        | Q(answer='N') & Q(question__no_score__gt=0)
    )
    return key_policy_answers.count(), key_policy_questions.count()


def get_and_split_document_questions(assessment):
    """
    Given the document questions group them by category, environmental, social
    and governance.
    Also map child questions and trigger questions.
    Trigger questions appear when the parent question is answered yes
    and child questions appear where the parent question is not answerable.
    """
    questions = DocumentQuestion.objects.filter(trigger_question__isnull=True).select_related(
        'company'
    ).select_related(
        'trigger_question'
    ).prefetch_related('documentquestion_set').prefetch_related('documentquestionanswerset_set').order_by('display_order')
    environmental_questions = []
    social_questions = []
    governance_questions = []
    for question in questions:
        is_required = question.is_required
        answered = False
        if question.is_answerable:
            answer = get_document_question_answer(assessment, question)
            if answer:
                if answer.answer:
                    question.answer = answer.answer
                    answered = True
                if answer.notes:
                    question.notes = answer.notes
        trigger_questions = []
        child_questions = []
        # trigger/child questions.
        for sub_question in question.documentquestion_set.all().order_by('display_order'):
            answer = get_document_question_answer(assessment, sub_question)
            if sub_question.is_required:
                is_required = True
            if answer and answer.answer:
                sub_question.answer = answer.answer
                answered = True
            if question.is_answerable:
                trigger_questions.append(sub_question)
            else:
                child_questions.append(sub_question)

        obj = {
            'question': question,
            'child_questions': child_questions,
            'trigger_questions': trigger_questions,
            'is_required': is_required,
            'answered': answered
        }
        if question.category == 'ENVIRONMENT':
            environmental_questions.append(obj)
        elif question.category == 'SOCIAL':
            social_questions.append(obj)
        elif question.category == 'GOVERNANCE':
            governance_questions.append(obj)

    return environmental_questions, social_questions, governance_questions


class CompanyVersionListView(LoginRequiredMixin, View):
    template_name = "assessment_version_list.html"

    def get(self, request, slug):
        company = self.request.user.userprofile.company
        assessment = get_assessment(slug, company)
        assessments = CompanyAssessment.objects.filter(
            parent=assessment
        ).order_by('-date_created')
        return render(request, self.template_name, locals())


class CompanyAssessmentKeyDetailsView(LoginRequiredMixin, View):
    template_name = "assessment_key_details.html"

    def get(self, request, slug=None):
        company = self.request.user.userprofile.company
        assessment = None
        if slug:
            assessment = get_assessment(slug, company)
            form = KeyDetailsForm(instance=assessment)
        else:
            assessment_count = CompanyAssessment.objects.filter(
                created_by__userprofile__company=company,
                is_latest=True
            ).count() + 1

            form = KeyDetailsForm(initial={
                'created_by': self.request.user,
                'report_reference': '{}/{}'.format(company.reference, leading_zeros(assessment_count, 4))
            })

        return render(request, self.template_name, {
            'assessment': assessment,
            'form': form,
            'previous_url': request.META.get('HTTP_REFERER')
        })

    def post(self, request, slug=None, *args, **kwargs):
        initial = {'created_by': self.request.user}
        assessment = None
        if slug:
            report_reference = request.POST.get('report_reference')
            assessment = get_assessment(
                slug,
                self.request.user.userprofile.company,
            )
            # assessment = create_new_version(assessment)

            form = KeyDetailsForm(request.POST, instance=assessment, initial=initial)
        else:
            form = KeyDetailsForm(request.POST, initial=initial)

        if form.is_valid():
            instance = form.save()

            instance.report_reference_slug = slugify(instance.report_reference)
            instance.save()

            sasb_industry = sasb_get_industry(instance)
            if sasb_industry:
                instance.industries.add(sasb_industry)
                material_risk_list = get_industry_risks(sasb_industry)

                for risk in material_risk_list:
                    instance.material_risks.add(risk)

            if is_ajax(request=request):
                return JsonResponse({
                    'footprintLink': request.build_absolute_uri(reverse(
                        'company_assessment_company_footprint',
                        kwargs={'slug': instance.slug}
                    ))
                })
            else:
                return redirect(
                    reverse(
                        'company_assessment_industries',
                        kwargs={'slug': instance.slug}
                    )
                )
        else:
            if is_ajax(request=request):
                return JsonResponse({
                    'status': 'FAIL',
                    'errors': form._errors
                })

        return render(request, self.template_name, {
            'assessment': assessment,
            'form': form,
            'previous_url': request.META.get('HTTP_REFERER')
        })


class CompanyIndustriesView(LoginRequiredMixin, View):
    template_name = "assessment_industries.html"

    def get(self, request, slug=None):
        company = self.request.user.userprofile.company
        assessment = None
        if slug:
            assessment = get_assessment(slug, company)
            form = CompanyIndustryForm(instance=assessment)

        total_count, industry_types = get_industry_types(assessment)

        return render(request, self.template_name, {
            'assessment': assessment,
            'form': form,
            'sasb_industry': sasb_get_industry(assessment),
            'total_count': total_count,
            'industry_types': industry_types
        })


class CompanyAssessmentCompanyFootprintView(LoginRequiredMixin, View):
    template_name = "assessment_company_footprint.html"

    def can_edit_score(self, permission):
        if permission and permission in self.request.user.userprofile.company.permissions.all():
            return True, Threshold.objects.all().values_list('text', flat=True)
        return False, []

    def get_risks_associated_with_industry(self, industries):
        material_risks_list = []

        for industry in industries:
            if industry.industryriskdataset_set.first() and industry.industryriskdataset_set.first().dataset_version:
                risks = industry.industryriskdataset_set.first().dataset_version.materialityrisk_set.all().order_by('risk_factor__name')
                for risk in risks:
                    if risk.risk_factor.slug not in material_risks_list:
                        material_risks_list.append(risk.risk_factor.slug)

        return material_risks_list

    def get(self, request, slug=None):
        assessment = get_assessment(slug, self.request.user.userprofile.company)
        form = CompanyFootprintForm(instance=assessment)
        material_risks_list = []
        allow_score_edit, score_options = self.can_edit_score(Permission.objects.filter(name='edit_risk_score').first())

        if hasattr(assessment, 'material_risks'):
            material_risks_list = list(assessment.material_risks.all().values_list('slug', flat=True))

        if assessment.industries:
            industry_data = list(assessment.industries.all())
            sasb_industry = sasb_get_industry(assessment)
            if sasb_industry and sasb_industry not in industry_data:
                industry_data = industry_data + [sasb_industry]

            # Each industry has a set of related risk factors if its blank we need to fetch.
            if not material_risks_list:
                material_risks_list = self.get_risks_associated_with_industry(industry_data)

            industry_data = [record.slug for record in industry_data]
            industry_data = ", ".join(industry_data)

        total_count, industry_types = get_industry_types(assessment)

        material_risks = ", ".join(material_risks_list)

        return render(request, self.template_name, locals())

    def post(self, request, slug=None, *args, **kwargs):
        assessment = get_assessment(slug, self.request.user.userprofile.company)
        # assessment = create_new_version(assessment)
        form = CompanyFootprintForm(request.POST, instance=assessment)

        if form.is_valid():
            instance = form.save()
            return redirect(
                reverse(
                    'company_management_questions',
                    kwargs={'slug': instance.slug}
                )
            )

        return render(request, self.template_name, locals())


class CompanyAssessmentDocumentCheckListView(LoginRequiredMixin, View):
    template_name = "assessment_document_checklist.html"

    def answer_trigger_question(self, assessment, question):
        parent_answer = get_document_question_answer(assessment, question.trigger_question)
        update_create = True
        if parent_answer and parent_answer.answer in ['N', 'X']:
            DocumentQuestionAnswerSet.objects.filter(
                assessment=assessment,
                question=question
            ).delete()
            update_create = False
        return update_create

    def get(self, request, slug=None):
        assessment = get_assessment(slug, self.request.user.userprofile.company)

        answers = ['Y', 'N', 'X']
        environmental_questions, social_questions, governance_questions = get_and_split_document_questions(assessment)

        answer_count = get_key_document_answer_count(assessment)

        return render(request, self.template_name, locals())

    def post(self, request, slug=None):
        assessment = get_assessment(slug, self.request.user.userprofile.company)
        # assessment = create_new_version(assessment)
        obj = {}
        answered_optional = False
        if is_ajax(request=request) and "file" in request.FILES:
            failed_uploads = []
            try:
                imported_data = tablib.Dataset().load(request.FILES['file'])
            except:
                return JsonResponse({
                    'status': 'FAIL'
                })
            for row in imported_data:
                if row[1]:
                    question = row[1]
                else:
                    question = None

                if question:
                    document_question = DocumentQuestion.objects.filter(text=question.strip())
                    if document_question:
                        document_question = document_question.first()
                        notes = row[5]
                        update_create = True
                        if document_question.trigger_question:
                            update_create = self.answer_trigger_question(assessment, document_question)
                        if update_create:
                            record, _ = DocumentQuestionAnswerSet.objects.get_or_create(
                                assessment=assessment,
                                question=document_question
                            )
                            if notes:
                                note_max_length = DocumentQuestionAnswerSet._meta.get_field('notes').max_length
                                record.notes = notes

                            if row[2]:
                                record.answer = 'Y'
                            elif row[3]:
                                record.answer = 'N'
                            else:
                                record.answer = 'X'

                            if document_question.is_required is False:
                                answered_optional = True

                            record.save()
                    else:
                        failed_uploads.append(question)
            if failed_uploads:
                return JsonResponse({
                    'status': 'partial-upload',
                    'failed': failed_uploads
                })

            assessment.answered_optional = answered_optional
            assessment.save()
            return JsonResponse({
                'status': 'OK'
            })
        else:
            form_items = dict(request.POST.items())
            cleaned_list = {}

            for key, value in form_items.items():
                if key not in ['csrfmiddlewaretoken', 'Next']:
                    if 'Notes' in key:
                        dict_key = key.replace('Notes', '')
                        if dict_key in cleaned_list.keys():
                            cleaned_list[dict_key] = cleaned_list[dict_key] + [{'notes': value}]
                        else:
                            cleaned_list[dict_key] = [{'notes': value}]
                    else:
                        if key in cleaned_list.keys():
                            cleaned_list[key] = cleaned_list[key] + [{'answer': value}]
                        else:
                            cleaned_list[key] = [{'answer': value}]

            cleared_questions = DocumentQuestionAnswerSet.objects.filter(assessment=assessment)
            for key, value in cleaned_list.items():
                user_response = dict(ChainMap(*value))
                if user_response.get('answer'):
                    question = DocumentQuestion.objects.get(id=key)
                    update_create = True
                    answer = user_response.get('answer')
                    if question.trigger_question:
                        update_create = self.answer_trigger_question(assessment, question)
                    if update_create:
                        record, _ = DocumentQuestionAnswerSet.objects.get_or_create(
                            assessment=assessment,
                            question=question
                        )
                        cleared_questions = cleared_questions.exclude(question=question)
                        if question.is_required is False:
                            answered_optional = True
                        note_max_length = DocumentQuestionAnswerSet._meta.get_field('notes').max_length
                        if user_response.get('notes'):
                            record.notes = user_response.get('notes')[:note_max_length]
                        else:
                            record.notes = None
                        record.answer = answer
                        record.save()
            cleared_questions.delete()
            assessment.answered_optional = answered_optional
            assessment.save()

        if is_ajax(request=request) and "file" not in request.FILES:
            rev_score = calculate_rev_counter_score(assessment)
            obj = {
                'revCounter': rev_score,
                'pin': rev_score / 100 * 180,
                'percentageComplete': calculate_document_question_complete(assessment)
            }

        if obj:
            return JsonResponse(obj)

        return redirect(
            reverse(
                'company_recommendations',
                kwargs={'slug': slug}
            )
        )


def create_question_answer(
    answer, assessment, insufficient_answer, priority_answer, question_text=None, question_id=None, company=None
):
    if company:
        question = ManagementQuestion.objects.filter(company=company, text=question_text).first()
        if not question:
            question = ManagementQuestion.objects.filter(text=question_text).first()
    elif question_text:
        question = ManagementQuestion.objects.get(text=question_text)
    elif question_id:
        question = ManagementQuestion.objects.get(id=question_id)

    if question:
        management_question_answer_obj, _ = ManagementQuestionAnswerSet.objects.get_or_create(
            assessment=assessment,
            question=question
        )
        insufficient = False
        if insufficient_answer.lower() == 'yes':
            insufficient = True

        priority = False
        if priority_answer.lower() == 'yes':
            priority = True

        management_question_answer_obj.answer = answer
        management_question_answer_obj.insufficient = insufficient
        management_question_answer_obj.priority = priority
        management_question_answer_obj.save()

        return question, management_question_answer_obj
    else:
        return None, None


class CompanyAssessmentManagementView(LoginRequiredMixin, View):
    template_name = "assessment_management.html"

    def get(self, request, slug=None):
        assessment = get_assessment(slug, self.request.user.userprofile.company)

        company = self.request.user.userprofile.company
        company_names = Company.objects.all().exclude(
            id=company.id
        ).select_related('country').select_related('primary_contact').select_related('account_manager').values_list('name', flat=True)

        industry_data = assessment.industries.all().select_related('type').values_list('slug', flat=True)

        top_5_impacts, overall_inherent_risk_score = calculate_top_5_impact_score(
            assessment.material_risks.all(), industry_data, assessment.countries.all().values_list('id', flat=True)
        )
        risk_slugs = [risk.slug for risk in top_5_impacts]

        environmental_questions, environmental_questions_ids = assessment.get_category_questions(company, 'ENVIRONMENT', risk_slugs)
        social_questions, social_question_ids = assessment.get_category_questions(company, 'SOCIAL', risk_slugs)
        governance_questions, governance_questions_ids = assessment.get_category_questions(company, 'GOVERNANCE', risk_slugs)

        exclude_risk_ids = environmental_questions_ids + social_question_ids + governance_questions_ids

        all_questions = environmental_questions + social_questions + governance_questions
        all_question_text = [question.text for question in all_questions]

        add_question_list = ManagementQuestion.objects.all().exclude(id__in=exclude_risk_ids).exclude(
            text__in=all_question_text
        ).exclude(company__name__in=company_names)

        return render(request, self.template_name, locals())

    def post(self, request, slug=None):
        assessment = get_assessment(slug, self.request.user.userprofile.company)
        # assessment = create_new_version(assessment)
        company = request.POST.get('company_id')
        question = request.POST.get('question')
        if company:
            company = get_object_or_404(Company, id=company)
        else:
            company = self.request.user.userprofile.company

        if company and question:
            question = get_object_or_404(ManagementQuestion, id=question)
            if question in assessment.deleted_management_questions.all():
                assessment.deleted_management_questions.remove(question)
                assessment.save()

            assessment.added_management_questions.add(question)
            assessment.save()
            redirect_url = reverse(
                'company_management_questions',
                kwargs={'slug': assessment.slug}
            )
            return redirect(redirect_url)
        elif "file" in request.FILES:
            try:
                imported_data = tablib.Dataset().load(request.FILES['file'])
            except:
                return JsonResponse({
                    'status': 'FAIL'
                })

            failed_uploads = []
            for row in imported_data:
                try:
                    answer = row[2]
                except IndexError:
                    return JsonResponse({
                        'status': 'FAIL'
                    })

                if row[3] and row[3].lower() == 'yes':
                    insufficient = 'Yes'
                else:
                    insufficient = 'No'

                if row[4] and row[4].lower() == 'yes':
                    priority = 'Yes'
                else:
                    priority = 'No'

                question, answer = create_question_answer(
                    answer=answer,
                    assessment=assessment,
                    insufficient_answer=insufficient,
                    priority_answer=priority,
                    question_text=row[1],
                    company=company
                )
                if not question:
                    failed_uploads.append(row[1])

            if failed_uploads:
                return JsonResponse({
                    'status': 'partial-upload',
                    'failed': failed_uploads
                })
            return JsonResponse({
                'status': 'OK'
            })
        else:
            form_items = json.loads(request.POST.get('form'))
            assessment.references = request.POST.get('references')
            assessment.save()
            for item in form_items:
                answer = item['value'][0]['value']
                if not answer:
                    answer = ''

                insufficient_answer = 'no'
                if item['value'][1]['value']:
                    insufficient_answer = 'yes'

                priority_answer = 'no'
                if item['value'][2]['value']:
                    priority_answer = 'yes'

                create_question_answer(
                    question_id=item['value'][0]['name'],
                    answer=answer,
                    assessment=assessment,
                    insufficient_answer=insufficient_answer,
                    priority_answer=priority_answer
                )

            return JsonResponse({
                'status': 'OK'
            })


class CompanyAssessmentRecommendationsView(LoginRequiredMixin, View):
    template_name = "assessment_recommendations.html"

    def get_risk_summary_text(self, threshold_text, company_name):
        return """Based on the available information reviewed (see the References), the Company would be assessed as {} risk with respect to ESG matters. This reports presents an initial screen of potential ESG issues associated with {} based on its industry type and countries of operation; there are likely further specific issues that will be identified by more detailed assessment of its operations.""".format(
            threshold_text,
            company_name,
        )

    def get(self, request, slug=None):
        assessment = get_assessment(slug, self.request.user.userprofile.company)
        industry_data = [industry.slug for industry in assessment.industries.all()]
        country_list = [country.id for country in assessment.countries.all()]

        top_5_impacts, overall_inherent_risk_score = calculate_top_5_impact_score(assessment.material_risks.all(), industry_data, country_list)
        issues = [risk.slug for risk in top_5_impacts]

        if assessment.custom_risk_score:
            threshold_text = assessment.custom_risk_score.text_in_sentence
        else:
            threshold_text = Threshold.objects.filter(lower_bound__lte=overall_inherent_risk_score, upper_bound__gte=overall_inherent_risk_score)
            if threshold_text:
                threshold_text = threshold_text.first().text_in_sentence

        if assessment.risk_summary and assessment.issues.all():
            issues = [risk.slug for risk in top_5_impacts] + [issue.slug for issue in assessment.issues.all()]

        form = RecommendationForm(
            instance=assessment,
            initial={
                'risk_summary': self.get_risk_summary_text(threshold_text, assessment.company_name)
            }
        )

        industry_data = ",".join(industry_data)
        issues = ",".join(issues)

        return render(request, self.template_name, locals())

    def post(self, request, slug=None, *args, **kwargs):
        assessment = get_assessment(slug, self.request.user.userprofile.company)
        # assessment = create_new_version(assessment)
        form = RecommendationForm(
            request.POST,
            instance=assessment,
            initial={'status': 'C'}
        )

        if form.is_valid():
            form.save()

            form_items = dict(request.POST.items())
            assessment.status = 'C'
            assessment.save()

            for key, value in form_items.items():
                if '-recommendation' in key:
                    issue, _ = AssessmentIssueRecommendations.objects.get_or_create(
                        assessment=assessment,
                        issue=key[:-len('-recommendation')]
                    )
                    if len(value) > 0:
                        issue.recommendation = value
                        issue.save()
                    else:
                        issue.delete()
            return redirect(
                reverse('dashboard')
            )

        return render(request, self.template_name, locals())


def Assessment_phase_redirect(request, slug):
    assessment = get_object_or_404(CompanyAssessment, slug=slug)

    if assessment.issues.all():
        return redirect(
            reverse(
                'company_recommendations',
                kwargs={'slug': assessment.slug}
            )
        )

    if assessment.documentquestionanswerset_set.all():
        return redirect(
            reverse(
                'company_management_questions',
                kwargs={'slug': assessment.slug}
            )
        )

    if assessment.material_risks.all():
        return redirect(
            reverse(
                'company_assessment_document_checklist',
                kwargs={'slug': assessment.slug}
            )
        )

    if assessment.company_name and not assessment.material_risks.all():
        return redirect(
            reverse(
                'company_assessment_company_footprint',
                kwargs={'slug': assessment.slug}
            )
        )

    return redirect(
        reverse(
            'company_assessment_key_detail',
            kwargs={'slug': assessment.slug}
        )
    )


def Industry_table(request):
    selected_industries = request.POST.getlist('industry[]')
    selected_risks = request.POST.getlist('risks[]')
    table_header_data = []
    risk_version = []
    risk_list = {}

    selected_risks = [risk for risk in selected_risks]

    if not selected_risks:
        assessment_slug = request.POST.get('assessment')
        assessment = CompanyAssessment.objects.get(slug=assessment_slug)
        selected_risks = list(assessment.material_risks.all().values_list('id', flat=True))

    records = IndustryRiskDataSet.objects.filter(
        industry__id__in=[industry for industry in selected_industries]
    )

    for record in records:
        if record.industry.name not in table_header_data:
            table_header_data.append(record.industry.name)

        if record.dataset_version:
            risk_version.append(record.dataset_version.id)

    table_header_data = table_header_data[:5]

    dataset_risks = RiskDataSet.objects.filter(id__in=selected_risks).order_by('name')

    material_risks = MaterialityRisk.objects.filter(
        dataset_version__id__in=risk_version,
        risk_factor__id__in=selected_risks
    )

    for risk in dataset_risks:
        risk_list[risk.name] = [
            risk.name,
            risk.description
        ]

    for header in table_header_data:
        material_risks_in_header = [material for material in material_risks if material.industry.name == header]
        for risk in dataset_risks:
            added = False
            if material_risks:
                material_risk = [material for material in material_risks_in_header if material.risk_factor.id == risk.id]
                if material_risk:
                    added = True
                    risk_list[risk.name] = risk_list[risk.name] + [{
                        'score': material_risk[0].materiality,
                        'industry': header,
                        'risk_colour': get_score_colour(material_risk[0].materiality)
                    }]

            if not added:
                risk_list[risk.name] = risk_list[risk.name] + [{
                    'score': None,
                    'industry': header,
                }]

    return JsonResponse({
        't_head_data': table_header_data,
        't_body_data': risk_list
    })


def Geographic_table(request):
    selected_countries = request.POST.getlist('countries[]')
    countries = Country.objects.filter(id__in=selected_countries).order_by('name')

    selected_risks = request.POST.getlist('risks[]')

    if not selected_risks:
        assessment_slug = request.POST.get('assessment')
        assessment = CompanyAssessment.objects.get(slug=assessment_slug)
        selected_risks = list(assessment.material_risks.all().values_list('id', flat=True))

    records = RiskDataSet.objects.filter(id__in=selected_risks).order_by('name')

    risk_list = {}

    if len(countries) > 5:
        active_version_list = [record.active_version.id for record in records if record.active_version]
        countries = calculate_top_5_at_risk_countrys(countries, active_version_list)
        selected_countries = [int(country.obj_id) for country in countries]

    table_header_data = [country.name for country in countries]
    table_header_data = table_header_data[:5]

    for record in records:
        risk_list[record.name] = [record.name, record.description]
        if record.active_version:
            for item in record.active_version.countryrisk_set.filter(country__id__in=selected_countries).order_by('country__name')[:5]:
                country = item.country.name
                if country and country in table_header_data:
                    exposure = item.exposure

                    if exposure:
                        risk_list[record.name] = risk_list[record.name] + [{
                            'score': exposure,
                            'country': country,
                            'risk_colour': get_score_colour(exposure)
                        }]

    return JsonResponse({
        't_head_data': table_header_data,
        't_body_data': risk_list
    })


def Risk_description(request):
    selected_risk_id = request.POST['risk_id']
    risk = RiskDataSet.objects.filter(id=selected_risk_id)
    industries = request.POST.getlist('industry[]')
    source = 'Anthesis'

    material_risks = get_risk_by_source(risk, industries)
    if material_risks:
        source = material_risks[0]['source']

    if source.lower() == 'client':
        source = 'Anthesis'

    return JsonResponse({
        'name': risk[0].name,
        'short_description': risk[0].description if risk[0].description else "No description available",
        'source': 'Sourced by {}'.format(source)
    })


def get_risk_by_source(selected_risks, selected_industries=None):
    risk_list = []
    selected_industries = [industry.replace(' ', '') for industry in selected_industries]
    selected_risk_ids = [risk.id for risk in selected_risks]
    if selected_industries:
        industries = Industry.objects.filter(slug__in=selected_industries)
        if not industries:
            industries = Industry.objects.filter(id__in=selected_industries)
        added_list = []
        for industry in industries:
            if industry.industryriskdataset_set.first() and industry.industryriskdataset_set.first().dataset_version:
                risks = industry.industryriskdataset_set.first().dataset_version.materialityrisk_set.filter(
                    risk_factor__id__in=selected_risk_ids
                ).order_by('risk_factor__name')
                for risk in risks:
                    if risk.risk_factor_id not in added_list:
                        colour = get_category_color(risk.risk_factor.category)

                        risk_list.append(
                            {
                                'id': risk.risk_factor_id,
                                'text': "{},{}".format(colour, risk.risk_factor.name),
                                'source': risk.source.name,
                                'name': risk.risk_factor.name
                            }
                        )
                        selected_risk_ids.remove(risk.risk_factor_id)
                        added_list.append(risk.risk_factor_id)
        for risk in selected_risks:
            if risk.id in selected_risk_ids:
                risk_list.append({
                    'id': risk.id,
                    'text': risk.name,
                    'source': 'Client',
                    'name': risk.name
                })
    else:
        risk_list = [{'id': risk.id, 'text': risk.name, 'source': 'Client', 'name': risk.name} for risk in selected_risks]

    return risk_list


def initial_selected_risk_options(request):
    selected_industries = None
    risks = request.GET.get('selected_risks', None)
    selected_risks = None

    if request.GET.get('selected_industries'):
        if ']' not in request.GET.get('selected_industries'):
            selected_industries = request.GET.get('selected_industries').split(',')
        else:
            try:
                selected_industries = request.GET.get('selected_industries').strip('][').split(',')
            except:
                selected_industries = None

    if risks:
        if not isinstance(risks, list):
            risks = risks.split(',')
            risks = [risk.strip() for risk in risks]
        else:
            risks = request.GET.get('selected_risks')

        selected_risks = RiskDataSet.objects.filter(
            slug__in=risks
        )

    if selected_risks and selected_industries:
        return JsonResponse({
            'risks': get_risk_by_source(selected_risks, selected_industries)
        })
    else:
        return JsonResponse({
            'No data': []
        })


def risk_options(request):
    search_term = request.GET.get('p')
    exclude = request.GET.getlist('exclude[]')
    if search_term:
        risks = RiskDataSet.objects.filter(
            name__icontains=search_term
        )
    else:
        risks = RiskDataSet.objects.all()

    if exclude:
        risks = risks.exclude(id__in=[risk_id for risk_id in exclude])

    risk_list = [{'id': risk.id, 'text': risk.name} for risk in risks]

    return JsonResponse({
        'risks': risk_list
    })


def industry_options(request):
    search_term = request.GET.get('p')
    if search_term:
        industries = Industry.objects.filter(name__icontains=search_term)
    else:
        industries = Industry.objects.all()

    industries_list = []
    for industry in industries:
        if industry.short_description:
            industry_text = "{}_{}".format(industry.name, industry.short_description)
        else:
            industry_text = "{}_{}".format(industry.name, '')
        industries_list.append({
            'id': industry.id,
            'text': industry_text
        })

    return JsonResponse({
        'industries': industries_list
    })


def Industry_select(request):
    selected_industries = request.POST.getlist('industry[]')
    material_risk_list = []
    added_list = []

    for selected_industry in selected_industries:
        try:
            industry = Industry.objects.get(pk=selected_industry)
        except:
            industry = Industry.objects.get(name=selected_industry)

        if industry.industryriskdataset_set.first() and industry.industryriskdataset_set.first().dataset_version:
            risks = industry.industryriskdataset_set.first().dataset_version.materialityrisk_set.all().order_by('risk_factor__name')
            for risk in risks:
                if risk.risk_factor_id not in added_list:
                    colour = get_category_color(risk.risk_factor.category)
                    material_risk_list.append(
                        {
                            'id': risk.risk_factor_id,
                            'text': "{},{}".format(colour, risk.risk_factor.name),
                        }
                    )
                    added_list.append(risk.risk_factor_id)

    return JsonResponse({'risks': material_risk_list})


def management_csv(request, slug, optional):
    headers = [
        'Category',
        'Question',
        'Answer',
        'Insufficient (Yes/No)',
        'Priority (Yes/No)'
    ]
    assessment = get_object_or_404(CompanyAssessment, slug=slug)
    company = request.user.userprofile.company

    wb = Workbook()
    sheet = wb.active
    sheet.append(headers)

    my_red = styles.colors.Color(rgb='fce7e8')
    my_fill = styles.fills.PatternFill(patternType='solid', fgColor=my_red)
    industry_data = [industry.slug for industry in assessment.industries.all()]
    country_list = [country.id for country in assessment.countries.all()]
    top_5_impacts, overall_inherent_risk_score = calculate_top_5_impact_score(assessment.material_risks.all(), industry_data, country_list)
    risk_slugs = [risk.slug for risk in top_5_impacts]
    if optional == 'optional':
        required_only = False
    else:
        required_only = True
    environmental_questions, environmental_questions_ids = assessment.get_category_questions(company, 'ENVIRONMENT', risk_slugs, required_only)
    social_questions, social_question_ids = assessment.get_category_questions(company, 'SOCIAL', risk_slugs, required_only)
    governance_questions, governance_questions_ids = assessment.get_category_questions(company, 'GOVERNANCE', risk_slugs, required_only)

    for question in list(environmental_questions) + list(social_questions) + list(governance_questions):
        sheet.append([
            question.category,
            question.text,
            '',
            '',
            ''
        ])
        question_risks = list(question.riskfactors.values_list('slug', flat=True))
        if any(elem in question_risks for elem in risk_slugs):
            sheet.cell(column=2, row=sheet.max_row).fill = my_fill

    sheetname = "Management_Questions.xlsx"
    response = HttpResponse(
        save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename={}'.format(sheetname)
    return response


def compile_document_questions(questions, category, required_only):
    question_list = []
    if required_only is False:
        for question in questions:
            if question['child_questions']:
                question_list.append([
                    '',
                    '       ' + str(question['question']),
                    'N/A',
                    'N/A',
                    'N/A',
                    'N/A'
                ])
                for child in question['child_questions']:
                    question_list.append([
                        category,
                        str(child.text),
                        '',
                        '',
                        ''
                    ])
            else:
                question_list.append([
                    category,
                    str(question['question']),
                    '',
                    '',
                    ''
                ])
                if question['trigger_questions']:
                    for trigger in question['trigger_questions']:
                        question_list.append([
                            '',
                            '       ' + str(trigger.text),
                            '',
                            '',
                            ''
                        ])
    else:
        for question in questions:
            if question['child_questions']:
                if question['is_required']:
                    question_list.append([
                        category,
                        str(question['question']),
                        'N/A',
                        'N/A',
                        'N/A',
                        'N/A'
                    ])
                for child in question['child_questions']:
                    if child.is_required:
                        question_list.append([
                            '',
                            '       ' + str(child.text),
                            '',
                            '',
                            ''
                        ])
            else:
                if question['is_required']:
                    question_list.append([
                        category,
                        str(question['question']),
                        '',
                        '',
                        ''
                    ])
                if question['trigger_questions']:
                    for trigger in question['trigger_questions']:
                        if trigger.is_required:
                            question_list.append([
                                '',
                                '       ' + str(trigger.text),
                                '',
                                '',
                                ''
                            ])

    return question_list


def document_checklist_csv(request, slug, optional):
    headers = [
        'Category',
        'Question',
        'Yes',
        'No',
        'No Evidence',
        'Notes ({} characters maximum)'.format(DocumentQuestionAnswerSet._meta.get_field('notes').max_length)
    ]
    assessment = get_object_or_404(CompanyAssessment, slug=slug)
    environmental_questions, social_questions, governance_questions = get_and_split_document_questions(assessment)
    wb = Workbook()
    sheet = wb.active
    sheet.append(headers)
    if optional == 'optional':
        required_only = False
    else:
        required_only = True
    environmental = compile_document_questions(environmental_questions, 'Environmental', required_only)
    for record in environmental:
        sheet.append(record)

    social = compile_document_questions(social_questions, 'Social', required_only)
    for record in social:
        sheet.append(record)

    governance = compile_document_questions(governance_questions, 'Governance', required_only)
    for record in governance:
        sheet.append(record)

    sheetname = "Maturity_Questions.xlsx"
    response = HttpResponse(
        save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename={}'.format(sheetname)
    return response


class footprintCsv(LoginRequiredMixin, View):
    template_name = "assessment_company_footprint.html"

    def create_int_list(self, records):
        records = list(set(records.split(',')))
        return [record for record in records if record]

    def create_headers(self, industries, country_id_list):
        headers = []
        headers.append('Risk')

        if industries:
            headers.append('Industry')
            headers.append('Industry Score')

            for index, value in enumerate(industries):
                if index > 0:
                    headers.append(' ')
                    headers.append(' ')

        if country_id_list:
            headers.append('Country')
            headers.append('Country Score')

            for index, value in enumerate(country_id_list):
                if index > 0:
                    headers.append(' ')
                    headers.append(' ')

        return headers

    def post(self, request):
        risk_ids = self.create_int_list(request.POST.get('risk_id[]'))
        industry_ids_list = self.create_int_list(request.POST.get('industry_ids[]'))
        country_id_list = self.create_int_list(request.POST.get('country_ids[]'))

        risks = RiskDataSet.objects.filter(id__in=risk_ids).order_by('name')
        countries = Country.objects.filter(id__in=country_id_list).order_by('name')
        industries = IndustryRiskDataSet.objects.filter(
            industry__id__in=industry_ids_list
        ).order_by('industry__name')

        data = tablib.Dataset(
            headers=self.create_headers(industries, country_id_list)
        )

        for risk in risks:
            data_list = []
            data_list.append(risk.name)

            for industry in industries:
                if industry.dataset_version:
                    data_list.append(industry.industry.name)
                    score = 'N/A'
                    for industry_risk in industry.dataset_version.materialityrisk_set.filter(
                        risk_factor__id=risk.id
                    ):
                        if industry_risk.materiality is not None:
                            score = industry_risk.materiality
                            break

                    data_list.append(score)

            for country in countries:
                data_list.append(country.name)
                score = 'N/A'
                if risk.active_version:
                    country_risk = CountryRisk.objects.filter(
                        country=country,
                        version__id=risk.active_version.id
                    )
                    if country_risk and country_risk[0].exposure is not None:
                        score = country_risk[0].exposure

                data_list.append(score)

            data.append(data_list)

        response = HttpResponse(data.export('customxlsx'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Footprint_results.xlsx"'

        return response


def hide_management_question(request, slug):
    assessment = get_object_or_404(CompanyAssessment, slug=slug)

    question = get_object_or_404(ManagementQuestion, id=request.GET.get('question_id'))

    if question in assessment.added_management_questions.all():
        assessment.added_management_questions.remove(question)
        assessment.save()

    assessment.deleted_management_questions.add(question)
    assessment.save()
    return JsonResponse({
        'state': 'success',
        'question': {
            'id': question.id,
            'text': question.text
        }
    })


def risk_recommendations(request):
    recommendation_list = []
    risks = RiskDataSet.objects.filter(
        id__in=[risk for risk in request.GET.getlist('risks[]')]
    )
    assessment_id = request.GET.get('assessment')
    assessment = None
    if assessment_id:
        assessment = CompanyAssessment.objects.get(id=assessment_id)

    if assessment:
        for risk in risks:
            issue = AssessmentIssueRecommendations.objects.filter(
                assessment=assessment,
                issue=risk.name
            )

            if issue:
                recommendation_list.append({
                    'risk': issue[0].issue,
                    'reccomendation': issue[0].recommendation
                })
            else:
                recommendation_list.append({
                    'risk': risk.name,
                    'reccomendation': risk.recommendation
                })

    else:
        for risk in risks:
            recommendation_list.append({
                'risk': risk.name,
                'reccomendation': risk.recommendation
            })

    return JsonResponse({'recommendations': recommendation_list})


def inherent_risk(request):
    assessment_slug = request.GET.get('assessment')
    if assessment_slug:
        assessment = CompanyAssessment.objects.get(slug=assessment_slug)
        if assessment and assessment.custom_risk_score:
            return JsonResponse({
                'score': assessment.custom_risk_score.text,
                'colour': assessment.custom_risk_score.colour
            })

    risks = RiskDataSet.objects.filter(
        id__in=request.GET.getlist('risks[]')
    )
    country_list = list(Country.objects.filter(id__in=request.GET.getlist('countries[]')).values_list('id', flat=True))
    industry_slugs = list(Industry.objects.filter(id__in=request.GET.getlist('industry[]')).values_list('slug', flat=True))

    _, overall_inherent_risk_score = calculate_top_5_impact_score(risks, industry_slugs, country_list)

    threshold_text = Threshold.objects.filter(lower_bound__lte=overall_inherent_risk_score, upper_bound__gte=overall_inherent_risk_score)
    if threshold_text:
        return JsonResponse({
            'score': threshold_text.first().text,
            'colour': threshold_text.first().colour
        })
    else:
        return JsonResponse({'score': 'N/A'})


def assessment_delete(request, slug):
    assessment = CompanyAssessment.objects.get(
        slug=slug,
        created_by__userprofile__company=request.user.userprofile.company
    )
    messages.add_message(request, messages.SUCCESS, "{} deleted successfully".format(assessment.company_name))
    assessment.delete()
    return redirect(reverse('dashboard'))


def partial_save(request):
    assessment_slug = request.POST.get('assessment-slug')
    assessment = CompanyAssessment.objects.get(slug=assessment_slug)
    # assessment = create_new_version(assessment)

    references = request.POST.get('references')
    if references:
        assessment.references = request.POST.get('references')
        assessment.save()
        return JsonResponse({
            'state': 'success',
        })

    question = get_object_or_404(ManagementQuestion, id=request.POST.get('question_id'))

    answer, _ = ManagementQuestionAnswerSet.objects.get_or_create(
        assessment=assessment,
        question=question
    )

    user_answer = request.POST.get('answer')

    if user_answer:
        answer.answer = user_answer
    else:
        slider = request.POST.get('slider')
        checked = True
        if request.POST.get('checked') == 'false':
            checked = False

        if slider == 'priority':
            answer.priority = checked
        if slider == 'insufficient':
            answer.insufficient = checked

    answer.save()

    return JsonResponse({
        'state': 'success',
    })

def add_industry_to_assessment(industry, assessment, material_risk_list):
    assessment.industries.add(industry)
    for risk in material_risk_list:
        assessment.material_risks.add(risk)


def remove_industry_from_assessment(industry, assessment, material_risk_list):
    assessment.industries.remove(industry)
    protected_risks = []
    for other_industry in assessment.industries.all():
        protected_risks.extend(get_industry_risks(other_industry))
    protected_risks = list(protected_risks)
    for risk in material_risk_list:
        if risk not in protected_risks:
            assessment.material_risks.remove(risk)


def get_industry_risks(industry):
    material_risk_list = []
    if industry.industryriskdataset_set.first() and industry.industryriskdataset_set.first().dataset_version:
        risks = industry.industryriskdataset_set.first().dataset_version.materialityrisk_set.all().order_by('risk_factor__name')
        for risk in risks:
            if risk.risk_factor not in material_risk_list:
                material_risk_list.append(risk.risk_factor)

    return material_risk_list


def assessment_add_remove_industry(request):
    checked = request.GET.get('checked')
    assessment = CompanyAssessment.objects.get(slug=request.GET.get('assessment_slug'))

    if request.GET.get('industry_slug'):
        industry = Industry.objects.get(slug=request.GET.get('industry_slug'))
    elif request.GET.get('industry_id'):
        industry = Industry.objects.get(id=request.GET.get('industry_id'))
    else:
        industry = Industry.objects.get(name=request.GET.get('industry_name'))

    material_risk_list = get_industry_risks(industry)

    if not checked:
        if industry in assessment.industries.all():
            remove_industry_from_assessment(industry, assessment, material_risk_list)
        else:
            add_industry_to_assessment(industry, assessment, material_risk_list)
    elif checked == 'true':
        add_industry_to_assessment(industry, assessment, material_risk_list)
    elif checked == 'false':
        remove_industry_from_assessment(industry, assessment, material_risk_list)

    assessment.save()

    return JsonResponse({
        'state': checked,
        'obj': {
            'slug': industry.slug,
            'id': industry.id
        }
    })


def sasb_company_check(request):
    company_name = request.GET.get('company_name')
    sasb_companies = list(SasbCompanyLookup.objects.filter(name__icontains=company_name).values_list('name', flat=True)[:20])
    if sasb_companies:
        return JsonResponse({
            'companies': sasb_companies
        })

    return JsonResponse({
        'companies': "No results - please fill in manually"
    })


def custom_risk_score_select(request):
    assessment = CompanyAssessment.objects.get(slug=request.GET.get('assessment_slug'))
    score = request.GET.get('custom_score')

    if 'please' in score.lower():
        assessment.custom_risk_score = None
        assessment.save()
        industry_data = [industry.slug for industry in assessment.industries.all()]
        country_list = [country.id for country in assessment.countries.all()]

        _, overall_inherent_risk_score = calculate_top_5_impact_score(assessment.material_risks.all(), industry_data, country_list)
        threshold_text = Threshold.objects.filter(lower_bound__lte=overall_inherent_risk_score, upper_bound__gte=overall_inherent_risk_score)
        if threshold_text:
           return JsonResponse({
                'score': threshold_text.first().text,
                'colour': threshold_text.first().colour
            })

        return JsonResponse({})
    else:
        threshold_text = Threshold.objects.get(text=score)
        assessment.custom_risk_score = threshold_text
        assessment.save()

        if threshold_text:
            return JsonResponse({
                'score': threshold_text.text,
                'colour': threshold_text.colour
            })
