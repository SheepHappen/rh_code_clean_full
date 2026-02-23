import re

from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from tablib.formats._xlsx import XLSXFormat as baseXlsx
from tablib.formats import registry

from django import forms

from accounts.models import UserProfile
from company.models import CompanyEmailDomain
from core.models import Threshold


INVALID_TITLE_REGEX = re.compile(r'[\\*?:/\[\]]')

def safe_xlsx_sheet_title(s, replace="-"):
    return re.sub(INVALID_TITLE_REGEX, replace, s)[:31]

class XLSXFormat():
    title = 'customxlsx'

    def get_height_for_row(cls, sheet, row_number):
        row = list(sheet.rows)[row_number]
        height = 0
        for cell in row:
            if len(str(cell.value)) > 100:
                height = len(str(cell.value)) / 4

        return height

    def export_set(cls, dataset, freeze_panes=True, invalid_char_subst="-"):
        """Returns XLSX representation of Dataset.
        If dataset.title contains characters which are considered invalid for an XLSX file
        sheet name (http://www.excelcodex.com/2012/06/worksheets-naming-conventions/), they will
        be replaced with `invalid_char_subst`.
        """
        wb = Workbook()
        ws = wb.active

        ws.title = (
            safe_xlsx_sheet_title(dataset.title, invalid_char_subst)
            if dataset.title else 'Tablib Dataset'
        )

        cls.dset_sheet(dataset, ws, freeze_panes)

        # format cell alignment and make it wraptext
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(
                    wrapText=True,
                    vertical='center',
                    horizontal='center'
                )

        # format cell width
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            if length < 100:
                ws.column_dimensions[column_cells[0].column_letter].width = length + 5
            else:
               ws.column_dimensions[column_cells[0].column_letter].width = length / 5

        # format cell height.
        for i in range(0, ws.max_row):
            height = cls.get_height_for_row( ws, i)
            if height > 0:
                ws.row_dimensions[i + 1].height = height

        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    def dset_sheet(cls, dataset, ws, freeze_panes=True):
        return baseXlsx.dset_sheet(dataset, ws, freeze_panes)

registry.register('customxlsx', XLSXFormat())


def get_score_colour(score):
    threshold_text = Threshold.objects.filter(lower_bound__lte=score, upper_bound__gte=score)
    if threshold_text:
        return threshold_text.first().colour


def check_registration_allowed(company):
    if company.user_signup == 'X':
        raise forms.ValidationError("""new employees can only be added in the company admin area""")


def check_email_domain(email_domain):
    company_email = CompanyEmailDomain.objects.filter(domain=email_domain, enabled=True).first()
    if not company_email:
        raise forms.ValidationError("""Unknown email domain""")

    return company_email


def validate_email(email):
    email_domain = email.split('@')[1].strip()
    company_email = check_email_domain(email_domain)

    if UserProfile.objects.filter(company=company_email.company, user__email=email).count():
        raise forms.ValidationError("The given email has already been registered: please log in or use a different email address.")

    check_registration_allowed(company_email.company)

    return email


def get_category_color(category):
    try:
        category = category.lower()
    except:
        category = category.name.lower()

    if category == 'environmental':
        return 'risk-green'
    elif category == 'social':
        return 'risk-red'
    elif category == 'governance':
        return 'risk-blue'


progress_bar = {
    "key_details": {
        "list_class": "",
        "has_inner": False,
        "span_state": "inactive-text",
        "url_name": 'company_assessment_key_detail',
        'name': 'Key details'
    },
    "industries": {
        "list_class": "",
        "has_inner": False,
        "span_state": "inactive-text",
        "url_name": 'company_assessment_industries',
        'name': 'Industries'
    },
    "company_footprint": {
        "list_class": "",
        "has_inner": False,
        "span_state": "inactive-text",
        "url_name": 'company_assessment_company_footprint',
        'name': 'Company footprint'
    },
    "management_questions": {
        "list_class": "",
        "has_inner": False,
        "span_state": "inactive-text",
        "url_name": 'company_management_questions',
        'name': 'Management questions'
    },
    "esg_maturity_review": {
        "list_class": "",
        "has_inner": False,
        "span_state": "inactive-text",
        "url_name": 'company_assessment_document_checklist',
        'name': 'ESG maturity review'
    },
    "recommendations": {
        "list_class": "",
        "has_inner": False,
        "span_state": "inactive-text",
        "url_name": 'company_recommendations',
        'name': 'Recommendations'
    },
    "kpi": {
        "list_class": "",
        "has_inner": False,
        "span_state": "inactive-text",
        "name": "KPIs & actions",
        "url_name": 'company_kpi',
    },
}

def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'